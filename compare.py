import pandas as pd
import openpyxl
import PIL # NOTE HOLY SHIT do not remove this it breaks all the spreadsheet images otherwise GOD

import tkinter as tk
from tkinter import filedialog

from clrprint import clrprint, clrinput
import sys

import ctypes

import warnings
import yaml

import math
import os

# Compile with `python -m PyInstaller placement-manual-comparison.spec`
# or `python -m PyInstaller --clean placement-manual-comparison.spec`
# The exe is placed in ./dist

#region Command Line Tools
def clr_line():
	print(f"{''.ljust(os.get_terminal_size().columns - 1, ' ')}", end="\r")
#endregion

#region Configuration
def get_config(filename):
	cwd = os.path.abspath(os.path.dirname(sys.executable))
	cfg_path = os.path.join(cwd, filename)
	if not os.path.exists(cfg_path):
		cfg_path = filename
	if not os.path.exists(cfg_path):
		return None

	with open(cfg_path, 'r') as f:
		cfg = yaml.safe_load(f)
	return cfg
#endregion

#region Write to Text File
def write_to_txt(lines, filename = 'included.txt'):
	str = ""
	print("Writing shared items to text file...", end="\r")
	for line in lines:
		str += f"{line}\n"

	with open(filename, 'w') as f:
		f.write(str)
	
	print("Writing shared items to text file [DONE]")
	print(f"List of shared items saved to \"{filename}\"")
#endregion

#region Get Shared Items
def get_items(path):
	print("Loading placement manual...", end="\r")
	df = pd.read_excel(path)
	clrprint("Loading placement manual [DONE]", clr='g')

	keywords = set([])
	for index, row in df.iterrows():
		if type(row[0]) == float and math.isnan(row[0]):
			continue
	
		if not type(row[0]) == int:
			continue

		print(f"Getting items from spreadsheet ({round(index/len(df.index)*100)}%)", end="\r")
		keywords.add(row[0])
	clrprint("Getting items from spreadsheet [DONE]", clr='g')
	return keywords

def get_shared_items(items, xl_list_path):
	print("Loading excel item list...", end="\r")
	df = pd.read_excel(xl_list_path)
	clrprint("Loading excel item list [DONE]", clr='g')

	shared = set([])
	msg = ""
	
	row_count = len(df.index)

	for index, row in df.iterrows():
		#print(, f"((({str(row)}>)))")
		msg = f"Searching row {index} of {row_count} for {list(row)[0]}... [{round(index/row_count*100)}%]"
		if list(row)[0] in items:
			shared.add(list(row)[0])
		print(msg, end="\r")

	print(f"{''.ljust(len(msg), ' ')}\r", end="")
	clrprint("Searching excel list for shared items [DONE]", clr='g')
	return shared

def write_to_spreadsheet(spreadsheet, output, shared, column, color = "00FFFF00", clear=True, log = lambda _ : None):
	print("Writing to spreadsheet... (this may take a while)", end="\r")
	log("> Loading workbook")
	workbook = openpyxl.load_workbook(spreadsheet)
	sheet = workbook.active

	prev = False
	set_bottom = None
	bold = openpyxl.styles.Side(border_style='thin', color='000000')
	side = openpyxl.styles.Side(border_style='thin', color='dcdcdc')
	sheet.column_dimensions[openpyxl.utils.get_column_letter(column+1)].width = 2.5
	for row in sheet.iter_rows(min_row=1, min_col=1, max_col=column+1):
		if row[0].value != None and type(row[0].value) == int:
			if clear: row[column].value = ""
			if row[0].value in shared:
				row[column].fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
				top = bold
				if prev: top = None
				row[column].border = openpyxl.styles.Border(left=bold, right=bold, top=top)
				set_bottom = row[column]
				prev = True
			elif clear:
				row[column].fill = openpyxl.styles.PatternFill(start_color="00FFFFFF", end_color=color, fill_type="solid")
				row[column].border = openpyxl.styles.Border(left=side, right=side, top=side, bottom=side)
				if set_bottom:
					set_bottom.border = openpyxl.styles.Border(left=set_bottom.border.left, right=set_bottom.border.right, top=set_bottom.border.top, bottom=bold)
				set_bottom = None
				prev = False
	
	readable = output.replace("\\", "/")
	log(rf'> Saving workbook to {readable}')
	workbook.save(output)

	clr_line()
	clrprint("Writing to spreadsheet [DONE]", clr='g')
	print(f"Spreadsheet saved to \"{output}\"")
	return output
#endregion

def compare(cfg, xl_base_path, xl_list_path, log = lambda _ : None):
	log("> Getting items")
	items = get_items(xl_base_path)

	log("> Comparing items")
	shared_items = get_shared_items(items, xl_list_path)

	out = xl_base_path
	color = f"00{cfg['color'][1:]}"
	col_num = openpyxl.utils.column_index_from_string(cfg['column']) - 1
	return write_to_spreadsheet(xl_base_path, out, shared_items, col_num, color, True, log=log)

def main():
	print("\nConfiguration:\n Open the 'config.yml' file in notepad to make changes\n colors are in aRGB format")
	print("Instructions: \n1. Navigate to the directory containing the placement manual spreadsheet and select it\n2. Do the same for the spreadsheet containing the list of items\n3. If asked, specify the number of the column (A being 0, B being 1, etc.) to overwrite and press [Enter]\n")
	print("You might have to click on the window again after any step to regain focus")
	input("Press [Enter] to continue")
	print("")

	root = tk.Tk()
	root.withdraw()

	#region Get Files
	print("Select placement manual", end="\r")
	xl_base_path = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xls")])
	clr_line()

	if not os.path.isfile(xl_base_path):
		clrprint("[Error] The provided excel spreadsheet does not exist", clr='r')
		print("Exiting...")
		return

	clrprint(xl_base_path, clr='b')


	print("Select line lists", end="\r")
	xl_list_path = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xls")])
	clr_line()

	if not os.path.isfile(xl_list_path):
		clrprint("[Error] The provided excel spreadsheet does not exist", clr='r')
		print("Exiting...")
		return
	
	clrprint(xl_list_path, clr='b')
	#endregion

	ctypes.windll.user32.BringWindowToTop(hwnd)

	#region Get Config
	cfg = get_config('config.yml')

	if cfg == None:
		clrprint("[Error] No configuration file found", clr='r')
		return

	if cfg['all']['suppress_python_warnings']:
		warnings.simplefilter("ignore")

	if cfg['compare']['ask_target_column']:
		target_col = input(f"Enter the target column (default is {openpyxl.utils.get_column_letter(cfg['compare']['target_column'] + 1)}) ")
		if target_col != "":
			cfg['compare']['target_column'] = openpyxl.utils.cell.column_index_from_string(target_col) - 1
	#endregion
	
	final_path = compare(cfg['compare'], xl_base_path, xl_list_path)
	clrinput('\nProcess completed, press [ENTER]', clr='g')

	owd = cfg['compare']['open_when_done']
	if type(owd) == bool:
		if owd: os.startfile(xl_base_path)
	else:
		op = input("Open placement manual [Y/n] ")
		if op.lower() == 'y'  or op == '':
			os.startfile(final_path)


if __name__ == "__main__":
	try:
		hwnd = ctypes.windll.user32.GetForegroundWindow()
		print("Welcome to Excel Search!\nPress Ctrl+C to quit at any time")
		#print("Enter a subcommand")
		#input("> ")

		main()
	except KeyboardInterrupt:
		clr_line()
		print("Program terminated by user")
